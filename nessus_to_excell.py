#!/usr/bin/env python3
"""
Nessus to Excel Converter
Converts Nessus XML scan results to professional Excel reports.

Copyright (C) 2025 Cuma KURT

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.

Developed by: Cuma KURT
Email: cumakurt@gmail.com
LinkedIn: https://www.linkedin.com/in/cuma-kurt-34414917/
"""

import argparse
import sys
import xml.etree.ElementTree as ET
import re
from pathlib import Path
from collections import defaultdict, Counter
from typing import List, Dict, Optional, Tuple, Any
from functools import lru_cache

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from tqdm import tqdm
except ImportError as e:
    print(f"Error: Required library not installed. Please install: pip install pandas openpyxl tqdm")
    print(f"Missing: {e.name}")
    sys.exit(1)

# ============================================================================
# Configuration Constants
# ============================================================================

# Severity colors (BGR format for Excel)
SEVERITY_COLORS = {
    'Critical': 'FF0000',
    'High': 'FF6600',
    'Medium': 'FFCC00',
    'Low': '00CCFF',
    'Info': 'CCCCCC'
}

# Header colors
HEADER_COLOR = '366092'
HEADER_FONT_COLOR = 'FFFFFF'
TITLE_BG_COLOR = 'D9E1F2'

# Row colors
ROW_EVEN_COLOR = 'F2F2F2'
ROW_ODD_COLOR = 'FFFFFF'

# Affected hosts count color thresholds
HIGH_COUNT_COLOR = 'FFE6E6'  # >= 10 hosts
MEDIUM_COUNT_COLOR = 'FFF4E6'  # >= 5 hosts

# Chart settings
CHART_HEIGHT = 18
CHART_WIDTH = 20
CHART_TITLE_FONT_SIZE = 16
CHART_AXIS_FONT_SIZE = 12
CHART_LABEL_FONT_SIZE = 11

# Font sizes
TITLE_FONT_SIZE = 16
HEADER_FONT_SIZE = 12
SECTION_HEADER_FONT_SIZE = 13
DATA_FONT_SIZE = 11
SMALL_FONT_SIZE = 9

# Column widths
MAX_COLUMN_WIDTH = 80
AFFECTED_HOSTS_MIN_WIDTH = 50
AFFECTED_HOSTS_COUNT_WIDTH = 18
VULNERABILITY_MAX_WIDTH = 40
DESCRIPTION_MAX_WIDTH = 60

# Severity order
SEVERITY_ORDER = {
    'Critical': 4,
    'High': 3,
    'Medium': 2,
    'Low': 1,
    'Info': 0
}

SEVERITY_LEVELS = ['Critical', 'High', 'Medium', 'Low', 'Info']

# Severity mapping (Nessus numeric to text)
SEVERITY_MAP = {
    '0': 'Info',
    '1': 'Low',
    '2': 'Medium',
    '3': 'High',
    '4': 'Critical'
}

# File size limits (in bytes)
MAX_FILE_SIZE = 500 * 1024 * 1024  # 500 MB

# Batch processing
BATCH_SIZE = 1000

# Top N values
TOP_HOSTS_COUNT = 10
TOP_PLUGINS_COUNT = 10


# ============================================================================
# Logging Configuration
# ============================================================================

import logging

def setup_logger(verbose: bool = False, log_file: Path = None) -> logging.Logger:
    """
    Setup and configure logger.
    
    Args:
        verbose: If True, set log level to DEBUG
        log_file: Optional path to log file
        
    Returns:
        Configured logger instance
    """
    logger = logging.getLogger('nessus_to_excel')
    logger.setLevel(logging.DEBUG if verbose else logging.INFO)
    
    # Remove existing handlers
    logger.handlers.clear()
    
    # Console handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.DEBUG if verbose else logging.INFO)
    console_format = logging.Formatter(
        '%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    console_handler.setFormatter(console_format)
    logger.addHandler(console_handler)
    
    # File handler (if specified)
    if log_file:
        log_file.parent.mkdir(parents=True, exist_ok=True)
        file_handler = logging.FileHandler(log_file)
        file_handler.setLevel(logging.DEBUG)
        file_format = logging.Formatter(
            '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        file_handler.setFormatter(file_format)
        logger.addHandler(file_handler)
    
    return logger


class NessusParserError(Exception):
    """Custom exception for Nessus parsing errors."""
    pass


class ExcelGenerationError(Exception):
    """Custom exception for Excel generation errors."""
    pass


class NessusParser:
    """Parser for Nessus XML files."""
    
    def __init__(self, logger: Optional[Any] = None, include_info: bool = False):
        """
        Initialize Nessus parser.
        
        Args:
            logger: Logger instance for logging
            include_info: If True, include Info (severity 0) vulnerabilities
        """
        self.vulnerabilities: List[Dict[str, Any]] = []
        self.host_vuln_count: Dict[str, int] = defaultdict(int)
        self.logger = logger
        self.processed_files: int = 0
        self.failed_files: List[Tuple[Path, str]] = []
        self.include_info = include_info
        
    def parse_file(self, file_path: Path) -> bool:
        """
        Parse a single .nessus file.
        
        Args:
            file_path: Path to .nessus file
            
        Returns:
            True if parsing successful, False otherwise
        """
        try:
            # Validate file
            if not self._validate_file(file_path):
                return False
            
            # Parse XML file
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            # Handle both NessusClientData_v2 and NessusClientData_v3 formats
            report = root.find('.//Report')
            if report is None:
                if self.logger:
                    self.logger.warning(f"No Report element found in {file_path}")
                self.failed_files.append((file_path, "No Report element found"))
                return False
            
            host_count = 0
            vuln_count_before = len(self.vulnerabilities)
            
            for report_host in report.findall('ReportHost'):
                host_count += 1
                host_name = report_host.get('name', 'Unknown')
                host_ip = self._extract_host_ip(report_host, host_name)
                
                # Process vulnerabilities
                for report_item in report_host.findall('ReportItem'):
                    self._process_vulnerability(report_item, host_name, host_ip)
            
            vuln_count_after = len(self.vulnerabilities)
            new_vulns = vuln_count_after - vuln_count_before
            
            self.processed_files += 1
            return True
            
        except ET.ParseError as e:
            self.failed_files.append((file_path, f"XML parsing error: {e}"))
            return False
        except Exception as e:
            self.failed_files.append((file_path, str(e)))
            return False
    
    def _validate_file(self, file_path: Path) -> bool:
        """
        Validate file before parsing.
        
        Args:
            file_path: Path to file
            
        Returns:
            True if valid, False otherwise
        """
        if not file_path.exists():
            return False
        
        if not file_path.is_file():
            return False
        
        # Check file size
        file_size = file_path.stat().st_size
        if file_size > MAX_FILE_SIZE:
            return False
        
        if file_size == 0:
            return False
        
        return True
    
    def _extract_host_ip(self, report_host: ET.Element, host_name: str) -> str:
        """
        Extract host IP from ReportHost element.
        
        Args:
            report_host: ReportHost XML element
            host_name: Default host name
            
        Returns:
            Host IP address
        """
        host_ip = None
        
        for host_prop in report_host.findall('HostProperties/tag'):
            prop_name = host_prop.get('name', '')
            if prop_name == 'host-ip':
                host_ip = host_prop.text
                break
            elif prop_name == 'host-fqdn' and not host_ip:
                host_ip = host_prop.text
        
        return host_ip if host_ip else host_name
    
    @staticmethod
    def _map_to_mitre_attack(
        plugin_name: str,
        description: str,
        synopsis: str,
        cve: str,
        port: str,
        service: str
    ) -> str:
        """
        Map vulnerability to MITRE ATT&CK techniques based on various indicators.
        
        Args:
            plugin_name: Plugin/vulnerability name
            description: Vulnerability description
            synopsis: Vulnerability synopsis
            cve: CVE identifier
            port: Port number
            service: Service name
            
        Returns:
            MITRE ATT&CK technique IDs (comma-separated if multiple)
        """
        techniques = set()
        text_lower = f"{plugin_name} {description} {synopsis}".lower()
        
        # SQL Injection
        if any(keyword in text_lower for keyword in ['sql injection', 'sqli', 'sql-injection', 'database injection']):
            techniques.add('T1190')  # Exploit Public-Facing Application
            techniques.add('T1059')  # Command and Scripting Interpreter
        
        # Cross-Site Scripting (XSS)
        if any(keyword in text_lower for keyword in ['xss', 'cross-site scripting', 'cross site scripting']):
            techniques.add('T1190')  # Exploit Public-Facing Application
            techniques.add('T1059.007')  # JavaScript
        
        # Remote Code Execution
        if any(keyword in text_lower for keyword in ['remote code execution', 'rce', 'code execution', 'arbitrary code']):
            techniques.add('T1190')  # Exploit Public-Facing Application
            techniques.add('T1059')  # Command and Scripting Interpreter
        
        # Command Injection
        if any(keyword in text_lower for keyword in ['command injection', 'os command injection', 'shell injection']):
            techniques.add('T1059')  # Command and Scripting Interpreter
            techniques.add('T1190')  # Exploit Public-Facing Application
        
        # Path Traversal
        if any(keyword in text_lower for keyword in ['path traversal', 'directory traversal', '../', '..\\']):
            techniques.add('T1190')  # Exploit Public-Facing Application
            techniques.add('T1005')  # Data from Local System
        
        # File Upload / File Inclusion
        if any(keyword in text_lower for keyword in ['file upload', 'file inclusion', 'arbitrary file', 'local file inclusion']):
            techniques.add('T1190')  # Exploit Public-Facing Application
            techniques.add('T1105')  # Ingress Tool Transfer
        
        # Authentication Bypass
        if any(keyword in text_lower for keyword in ['authentication bypass', 'auth bypass', 'login bypass', 'credential bypass']):
            techniques.add('T1078')  # Valid Accounts
            techniques.add('T1190')  # Exploit Public-Facing Application
        
        # Weak Credentials / Default Passwords
        if any(keyword in text_lower for keyword in ['weak password', 'default password', 'weak credentials', 'default credentials']):
            techniques.add('T1078')  # Valid Accounts
            techniques.add('T1110')  # Brute Force
        
        # Privilege Escalation
        if any(keyword in text_lower for keyword in ['privilege escalation', 'privilege elevation', 'sudo', 'suid']):
            techniques.add('T1068')  # Exploitation for Privilege Escalation
            techniques.add('T1078')  # Valid Accounts
        
        # Information Disclosure
        if any(keyword in text_lower for keyword in ['information disclosure', 'information leak', 'sensitive information', 'data exposure']):
            techniques.add('T1005')  # Data from Local System
            techniques.add('T1040')  # Network Sniffing
        
        # Denial of Service
        if any(keyword in text_lower for keyword in ['denial of service', 'dos', 'ddos', 'resource exhaustion']):
            techniques.add('T1499')  # Endpoint Denial of Service
            techniques.add('T1498')  # Network Denial of Service
        
        # Network Sniffing / Man-in-the-Middle
        if any(keyword in text_lower for keyword in ['sniffing', 'man-in-the-middle', 'mitm', 'ssl/tls', 'certificate']):
            techniques.add('T1040')  # Network Sniffing
            techniques.add('T1557')  # Adversary-in-the-Middle
        
        # SSRF (Server-Side Request Forgery)
        if any(keyword in text_lower for keyword in ['ssrf', 'server-side request forgery', 'request forgery']):
            techniques.add('T1190')  # Exploit Public-Facing Application
            techniques.add('T1105')  # Ingress Tool Transfer
        
        # XXE (XML External Entity)
        if any(keyword in text_lower for keyword in ['xxe', 'xml external entity', 'xml injection']):
            techniques.add('T1190')  # Exploit Public-Facing Application
            techniques.add('T1005')  # Data from Local System
        
        # Deserialization
        if any(keyword in text_lower for keyword in ['deserialization', 'unserialize', 'pickle', 'java deserialization']):
            techniques.add('T1190')  # Exploit Public-Facing Application
            techniques.add('T1059')  # Command and Scripting Interpreter
        
        # Port and Service based mapping
        if port and port != 'N/A':
            try:
                port_num = int(port)
                # SSH
                if port_num == 22:
                    if 'weak' in text_lower or 'default' in text_lower or 'brute' in text_lower:
                        techniques.add('T1078')  # Valid Accounts
                        techniques.add('T1110')  # Brute Force
                # FTP
                elif port_num == 21:
                    if 'anonymous' in text_lower or 'weak' in text_lower:
                        techniques.add('T1078')  # Valid Accounts
                        techniques.add('T1105')  # Ingress Tool Transfer
                # SMB
                elif port_num in [139, 445]:
                    if 'smb' in text_lower or 'samba' in text_lower:
                        techniques.add('T1021.002')  # SMB/Windows Admin Shares
                        techniques.add('T1078')  # Valid Accounts
                # RDP
                elif port_num == 3389:
                    if 'rdp' in text_lower or 'remote desktop' in text_lower:
                        techniques.add('T1021.001')  # Remote Desktop Protocol
                        techniques.add('T1078')  # Valid Accounts
                # Database ports
                elif port_num in [3306, 1433, 5432]:
                    if 'sql' in text_lower or 'database' in text_lower:
                        techniques.add('T1078')  # Valid Accounts
                        techniques.add('T1005')  # Data from Local System
            except (ValueError, TypeError):
                pass
        
        # Service-based mapping
        if service and service != 'N/A':
            service_lower = service.lower()
            if 'http' in service_lower or 'https' in service_lower:
                if 'injection' in text_lower or 'xss' in text_lower:
                    techniques.add('T1190')  # Exploit Public-Facing Application
            elif 'ssh' in service_lower:
                techniques.add('T1021.004')  # SSH
                techniques.add('T1078')  # Valid Accounts
            elif 'ftp' in service_lower:
                techniques.add('T1105')  # Ingress Tool Transfer
                techniques.add('T1078')  # Valid Accounts
        
        # If no specific technique found, use general exploitation technique
        if not techniques:
            techniques.add('T1190')  # Exploit Public-Facing Application
        
        return ', '.join(sorted(techniques))
    
    def _process_vulnerability(
        self, 
        report_item: ET.Element, 
        host_name: str, 
        host_ip: str
    ) -> None:
        """
        Process a single vulnerability item.
        
        Args:
            report_item: ReportItem XML element
            host_name: Host name
            host_ip: Host IP address
        """
        plugin_id = report_item.get('pluginID', 'N/A')
        plugin_name = report_item.get('pluginName', 'N/A')
        severity = report_item.get('severity', '0')
        
        # Convert severity number to text
        severity_text = SEVERITY_MAP.get(severity, 'Unknown')
        
        # Skip Info (severity 0) vulnerabilities unless include_info is True
        if not self.include_info and severity == '0':
            return
        
        # Extract additional information
        description = self._get_text(report_item, 'description', '')
        solution = self._get_text(report_item, 'solution', '')
        synopsis = self._get_text(report_item, 'synopsis', '')
        plugin_output = self._get_text(report_item, 'plugin_output', '')
        cvss_base_score = self._get_text(report_item, 'cvss_base_score', '')
        cvss_vector = self._get_text(report_item, 'cvss_vector', '')
        cve = self._get_text(report_item, 'cve', '')
        port = report_item.get('port', 'N/A')
        protocol = report_item.get('protocol', 'N/A')
        svc_name = report_item.get('svc_name', 'N/A')
        
        # Extract exploit information - check both attributes and child elements
        exploit_available = report_item.get('exploit_available', '')
        if not exploit_available:
            exploit_available = self._get_text(report_item, 'exploit_available', '')
        
        exploitability_ease = report_item.get('exploitability_ease', '')
        if not exploitability_ease:
            exploitability_ease = self._get_text(report_item, 'exploitability_ease', '')
        
        exploit_framework_canvas = report_item.get('exploit_framework_canvas', '')
        if not exploit_framework_canvas:
            exploit_framework_canvas = self._get_text(report_item, 'exploit_framework_canvas', '')
        
        exploit_framework_metasploit = report_item.get('exploit_framework_metasploit', '')
        if not exploit_framework_metasploit:
            exploit_framework_metasploit = self._get_text(report_item, 'exploit_framework_metasploit', '')
        
        exploit_framework_core = report_item.get('exploit_framework_core', '')
        if not exploit_framework_core:
            exploit_framework_core = self._get_text(report_item, 'exploit_framework_core', '')
        
        # Determine exploit status
        exploit_status = self._determine_exploit_status(
            exploit_available,
            exploitability_ease,
            exploit_framework_canvas,
            exploit_framework_metasploit,
            exploit_framework_core
        )
        
        # Determine exposure (Internal/External) based on RFC1918
        exposure = 'External' if not self._is_rfc1918_ip(host_ip) else 'Internal'
        
        # Calculate organization-specific risk score
        risk_score = self._calculate_risk_score(
            cvss_base_score,
            port,
            protocol,
            exploit_status,
            exposure
        )
        
        # Map to MITRE ATT&CK techniques
        mitre_attack = self._map_to_mitre_attack(
            plugin_name,
            description,
            synopsis,
            cve,
            port,
            svc_name
        )
        
        vuln_data = {
            'Host Name': host_name,
            'Host IP': host_ip,
            'Exposure': exposure,
            'Plugin ID': plugin_id,
            'Plugin Name': plugin_name,
            'Severity': severity_text,
            'Severity Value': int(severity),
            'Port': port,
            'Protocol': protocol,
            'Service': svc_name,
            'Synopsis': synopsis,
            'Description': description,
            'Solution': solution,
            'Plugin Output': plugin_output,
            'CVSS Base Score': cvss_base_score,
            'CVSS Vector': cvss_vector,
            'CVE': cve,
            'Risk Score': risk_score,
            'MITRE ATT&CK': mitre_attack,
            'Exploit Available': exploit_available,
            'Exploitability Ease': exploitability_ease,
            'Exploit Status': exploit_status,
            'Exploit Framework Canvas': exploit_framework_canvas,
            'Exploit Framework Metasploit': exploit_framework_metasploit,
            'Exploit Framework Core': exploit_framework_core
        }
        
        self.vulnerabilities.append(vuln_data)
        self.host_vuln_count[host_ip] += 1
    
    @staticmethod
    def _get_text(element: ET.Element, tag_name: str, default: str = '') -> str:
        """
        Safely get text content from an XML element.
        
        Args:
            element: XML element
            tag_name: Tag name to find
            default: Default value if not found
            
        Returns:
            Text content or default value
        """
        found = element.find(tag_name)
        return found.text if found is not None and found.text else default
    
    @staticmethod
    def _is_rfc1918_ip(ip_address: str) -> bool:
        """
        Check if IP address is in RFC1918 private address space.
        
        Args:
            ip_address: IP address string
            
        Returns:
            True if IP is in RFC1918 private range, False otherwise
        """
        try:
            parts = ip_address.split('.')
            if len(parts) != 4:
                return False
            
            first_octet = int(parts[0])
            second_octet = int(parts[1])
            
            # 10.0.0.0/8
            if first_octet == 10:
                return True
            
            # 172.16.0.0/12
            if first_octet == 172 and 16 <= second_octet <= 31:
                return True
            
            # 192.168.0.0/16
            if first_octet == 192 and second_octet == 168:
                return True
            
            return False
        except (ValueError, AttributeError):
            return False
    
    @staticmethod
    def _calculate_risk_score(
        cvss_score: str,
        port: str,
        protocol: str,
        exploit_status: str,
        exposure: str
    ) -> float:
        """
        Calculate organization-specific risk score based on CVSS, port, protocol, exploit status, and exposure.
        
        Args:
            cvss_score: CVSS base score
            port: Port number
            protocol: Protocol (tcp/udp)
            exploit_status: Exploit status
            exposure: Internal or External
            
        Returns:
            Risk score (0-100)
        """
        # Base CVSS score (0-10)
        try:
            base_score = float(cvss_score) if cvss_score and cvss_score != 'N/A' else 0.0
        except (ValueError, TypeError):
            base_score = 0.0
        
        # Port risk multiplier
        port_risk = 1.0
        if port and port != 'N/A':
            try:
                port_num = int(port)
                # High-risk ports
                if port_num in [21, 22, 23, 25, 53, 80, 110, 143, 443, 445, 1433, 3306, 3389, 5432, 8080]:
                    port_risk = 1.2
                # Very high-risk ports
                elif port_num in [135, 139, 445, 1433, 3306, 3389]:
                    port_risk = 1.3
            except (ValueError, TypeError):
                pass
        
        # Protocol risk
        protocol_risk = 1.1 if protocol and protocol.lower() == 'tcp' else 1.0
        
        # Exploit status multiplier
        exploit_multiplier = 1.0
        if exploit_status:
            exploit_lower = exploit_status.lower()
            if 'yes' in exploit_lower:
                exploit_multiplier = 1.5
            elif 'possibly' in exploit_lower:
                exploit_multiplier = 1.2
        
        # Exposure multiplier (External is higher risk)
        exposure_multiplier = 1.3 if exposure == 'External' else 1.0
        
        # Calculate risk score (0-100)
        risk_score = base_score * port_risk * protocol_risk * exploit_multiplier * exposure_multiplier
        
        # Normalize to 0-100 scale
        risk_score = min(risk_score * 10, 100.0)
        
        return round(risk_score, 2)
    
    @staticmethod
    def _determine_exploit_status(
        exploit_available: str,
        exploitability_ease: str,
        exploit_framework_canvas: str,
        exploit_framework_metasploit: str,
        exploit_framework_core: str
    ) -> str:
        """
        Determine exploit status based on available exploit information.
        
        Args:
            exploit_available: Exploit available flag
            exploitability_ease: Exploitability ease level
            exploit_framework_canvas: Canvas framework exploit
            exploit_framework_metasploit: Metasploit framework exploit
            exploit_framework_core: Core framework exploit
            
        Returns:
            Exploit status string
        """
        # Check if exploit is available
        if exploit_available and exploit_available.lower() in ['true', 'yes', '1']:
            # Check exploitability ease
            if exploitability_ease:
                ease_lower = exploitability_ease.lower()
                if 'exploits are available' in ease_lower or 'easy' in ease_lower:
                    return 'Yes - Easy'
                elif 'exploit' in ease_lower:
                    return 'Yes - Available'
            
            # Check framework exploits
            frameworks = []
            if exploit_framework_canvas and exploit_framework_canvas.lower() in ['true', 'yes', '1']:
                frameworks.append('Canvas')
            if exploit_framework_metasploit and exploit_framework_metasploit.lower() in ['true', 'yes', '1']:
                frameworks.append('Metasploit')
            if exploit_framework_core and exploit_framework_core.lower() in ['true', 'yes', '1']:
                frameworks.append('Core')
            
            if frameworks:
                return f'Yes - {", ".join(frameworks)}'
            
            return 'Yes'
        
        # Check if exploitability ease indicates availability
        if exploitability_ease:
            ease_lower = exploitability_ease.lower()
            if 'exploit' in ease_lower and 'not' not in ease_lower:
                return 'Possibly'
            elif 'no known' in ease_lower or 'not available' in ease_lower:
                return 'No'
        
        return 'Unknown'


class ExcelFormatHelper:
    """Helper class for Excel formatting operations."""
    
    @staticmethod
    def get_header_style() -> Tuple[PatternFill, Font]:
        """Get header cell style."""
        fill = PatternFill(
            start_color=HEADER_COLOR,
            end_color=HEADER_COLOR,
            fill_type="solid"
        )
        font = Font(
            bold=True,
            color=HEADER_FONT_COLOR,
            size=HEADER_FONT_SIZE
        )
        return fill, font
    
    @staticmethod
    def get_title_style() -> Tuple[PatternFill, Font]:
        """Get title cell style."""
        fill = PatternFill(
            start_color=TITLE_BG_COLOR,
            end_color=TITLE_BG_COLOR,
            fill_type="solid"
        )
        font = Font(bold=True, size=TITLE_FONT_SIZE, color=HEADER_COLOR)
        return fill, font
    
    @staticmethod
    def get_border() -> Border:
        """Get standard border style."""
        return Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    @staticmethod
    def get_row_fills() -> Tuple[PatternFill, PatternFill]:
        """Get even and odd row fill colors."""
        fill_even = PatternFill(
            start_color=ROW_EVEN_COLOR,
            end_color=ROW_EVEN_COLOR,
            fill_type="solid"
        )
        fill_odd = PatternFill(
            start_color=ROW_ODD_COLOR,
            end_color=ROW_ODD_COLOR,
            fill_type="solid"
        )
        return fill_even, fill_odd
    
    @staticmethod
    def get_severity_fill(severity: str) -> Optional[PatternFill]:
        """Get severity-specific fill color."""
        color = SEVERITY_COLORS.get(severity)
        if color:
            return PatternFill(
                start_color=color,
                end_color=color,
                fill_type="solid"
            )
        return None


class ExcelReportGenerator:
    """Generate professional Excel reports from parsed Nessus data."""
    
    def __init__(
        self, 
        vulnerabilities: List[Dict[str, Any]], 
        host_vuln_count: Dict[str, int],
        logger: Optional[Any] = None
    ):
        """
        Initialize Excel report generator.
        
        Args:
            vulnerabilities: List of vulnerability dictionaries
            host_vuln_count: Dictionary mapping host IPs to vulnerability counts
            logger: Logger instance
        """
        self.vulnerabilities = vulnerabilities
        self.host_vuln_count = host_vuln_count
        self.output_file: Optional[Path] = None
        self.severity_plugin_positions: Dict[str, Dict[str, int]] = {}
        self.logger = logger
        self.format_helper = ExcelFormatHelper()
    
    def generate(self, output_path: Path) -> None:
        """
        Generate Excel report with Summary and Vulnerabilities sheets.
        
        Args:
            output_path: Path to output Excel file
            
        Raises:
            ExcelGenerationError: If generation fails
        """
        try:
            self.output_file = Path(output_path)
            if self.logger:
                self.logger.info(f"Generating Excel report: {output_path}")
            
            # Create DataFrame from vulnerabilities
            df = pd.DataFrame(self.vulnerabilities)
            
            if len(df) == 0:
                raise ExcelGenerationError("No vulnerabilities to process")
            
            # Rename Plugin Name to Vulnerability and remove Plugin ID
            df = df.rename(columns={'Plugin Name': 'Vulnerability'})
            if 'Plugin ID' in df.columns:
                df = df.drop('Plugin ID', axis=1)
            
            # Sort by severity (Critical first) and then by host
            df['Severity Order'] = df['Severity'].map(SEVERITY_ORDER).fillna(0)
            df = df.sort_values(
                ['Severity Order', 'Host IP', 'Vulnerability'],
                ascending=[False, True, True]
            )
            df = df.drop('Severity Order', axis=1)
            
            # Write to Excel
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Write summary sheet first
                self._create_summary_sheet(writer)
                
                # Write CVE sheet
                self._create_cve_sheet(writer, df)
                
                # Write PCI-DSS sheet
                self._create_pci_dss_sheet(writer, df)
                
                # Write severity-specific sheets
                self._create_severity_sheets(writer, df)
                
                # Write all vulnerabilities sheet
                df.to_excel(writer, sheet_name='All Vulnerabilities', index=False)
            
            # Format the Excel file
            self._format_excel(output_path)
            
            if self.logger:
                self.logger.info(f"Successfully generated Excel report: {output_path}")
            
        except Exception as e:
            error_msg = f"Error generating Excel report: {e}"
            if self.logger:
                self.logger.error(error_msg, exc_info=True)
            raise ExcelGenerationError(error_msg) from e
    
    def _create_summary_sheet(self, writer: pd.ExcelWriter) -> None:
        """Create Summary sheet with statistics and charts."""
        summary_data = []
        
        # Total vulnerabilities
        total_vulns = len(self.vulnerabilities)
        summary_data.append(['Vulnerability Report Summary', ''])
        summary_data.append([])
        summary_data.append(['Total Vulnerabilities', total_vulns])
        summary_data.append(['Total Hosts Scanned', len(self.host_vuln_count)])
        
        # Calculate risk statistics
        risk_scores = [v.get('Risk Score', 0) for v in self.vulnerabilities if v.get('Risk Score')]
        if risk_scores:
            avg_risk = sum(risk_scores) / len(risk_scores)
            max_risk = max(risk_scores)
            high_risk_count = len([r for r in risk_scores if r >= 70])
            critical_risk_count = len([r for r in risk_scores if r >= 90])
        else:
            avg_risk = 0
            max_risk = 0
            high_risk_count = 0
            critical_risk_count = 0
        
        summary_data.append([])  # Empty row
        summary_data.append(['Risk Score Statistics', ''])
        summary_data.append(['Average Risk Score', f'{avg_risk:.2f}'])
        summary_data.append(['Maximum Risk Score', f'{max_risk:.2f}'])
        summary_data.append(['High Risk Vulnerabilities (>=70)', high_risk_count])
        summary_data.append(['Critical Risk Vulnerabilities (>=90)', critical_risk_count])
        summary_data.append([])  # Empty row
        
        # Exposure distribution
        exposure_counts = Counter(v.get('Exposure', 'Unknown') for v in self.vulnerabilities)
        summary_data.append(['Exposure Distribution', ''])
        summary_data.append(['Exposure', 'Count'])
        exposure_start_row = len(summary_data) + 1
        for exposure in ['Internal', 'External']:
            summary_data.append([exposure, exposure_counts.get(exposure, 0)])
        exposure_end_row = len(summary_data)
        summary_data.append([])  # Empty row
        
        # Severity distribution
        severity_counts = Counter(v['Severity'] for v in self.vulnerabilities)
        summary_data.append(['Severity Distribution', ''])
        summary_data.append(['Severity', 'Count'])
        
        severity_start_row = len(summary_data) + 1
        for severity in SEVERITY_LEVELS:
            summary_data.append([severity, severity_counts.get(severity, 0)])
        
        severity_end_row = len(summary_data)
        summary_data.append([])  # Empty row
        
        # Top N affected hosts
        summary_data.append([f'Top {TOP_HOSTS_COUNT} Most Affected Hosts', ''])
        summary_data.append(['Host IP', 'Vulnerability Count'])
        
        hosts_start_row = len(summary_data) + 1
        top_hosts = sorted(
            self.host_vuln_count.items(),
            key=lambda x: x[1],
            reverse=True
        )[:TOP_HOSTS_COUNT]
        
        for host_ip, count in top_hosts:
            summary_data.append([host_ip, count])
        
        hosts_end_row = len(summary_data)
        
        # Create DataFrame and write to Excel
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False, header=False)
        
        # Store row positions for chart creation
        self.summary_severity_start = severity_start_row
        self.summary_severity_end = severity_end_row
        self.summary_hosts_start = hosts_start_row
        self.summary_hosts_end = hosts_end_row
        self.summary_exposure_start = exposure_start_row
        self.summary_exposure_end = exposure_end_row
    
    def _normalize_cve(self, cve_string: str) -> List[str]:
        """
        Normalize and extract CVE IDs from a CVE string.
        
        Args:
            cve_string: CVE string that may contain multiple CVEs
            
        Returns:
            List of normalized CVE IDs
        """
        if not cve_string or cve_string == 'N/A' or cve_string.strip() == '':
            return []
        
        # Split by common delimiters
        cves = []
        for delimiter in [',', ';', '\n', '\r\n', '|']:
            if delimiter in cve_string:
                cves = [c.strip() for c in cve_string.split(delimiter)]
                break
        
        if not cves:
            cves = [cve_string.strip()]
        
        # Normalize CVE format (CVE-YYYY-NNNNN)
        normalized_cves = []
        for cve in cves:
            cve = cve.strip().upper()
            # Match CVE pattern
            if cve.startswith('CVE-'):
                # Extract CVE ID
                parts = cve.split('-')
                if len(parts) >= 3:
                    normalized_cves.append('-'.join(parts[:3]))
            elif 'CVE' in cve:
                # Try to extract CVE from text
                cve_pattern = r'CVE-\d{4}-\d{4,}'
                matches = re.findall(cve_pattern, cve)
                normalized_cves.extend(matches)
        
        return list(set(normalized_cves))  # Remove duplicates
    
    def _create_cve_sheet(self, writer: pd.ExcelWriter, df: pd.DataFrame) -> None:
        """Create CVE sheet with grouped vulnerabilities by CVE."""
        # Extract and normalize all CVEs
        cve_data = defaultdict(lambda: {
            'CVE': '',
            'Affected Hosts': [],
            'Vulnerabilities': [],
            'Severities': set(),
            'CVSS Scores': [],
            'Exploit Status': set(),
            'Total Occurrences': 0
        })
        
        for _, row in df.iterrows():
            cve_string = row.get('CVE', '')
            if not cve_string or cve_string == 'N/A':
                continue
            
            cves = self._normalize_cve(cve_string)
            if not cves:
                continue
            
            for cve in cves:
                host_info = f"{row['Host IP']}"
                if row.get('Host Name', '') and row['Host Name'] != row['Host IP']:
                    host_info += f" ({row['Host Name']})"
                
                if cve_data[cve]['CVE'] == '':
                    cve_data[cve]['CVE'] = cve
                
                cve_data[cve]['Affected Hosts'].append(host_info)
                cve_data[cve]['Vulnerabilities'].append(row.get('Vulnerability', 'N/A'))
                cve_data[cve]['Severities'].add(row.get('Severity', 'Unknown'))
                cvss_score = row.get('CVSS Base Score', '')
                if cvss_score and cvss_score != 'N/A':
                    try:
                        cve_data[cve]['CVSS Scores'].append(float(cvss_score))
                    except (ValueError, TypeError):
                        pass
                exploit_status = row.get('Exploit Status', 'Unknown')
                if exploit_status and exploit_status != 'Unknown':
                    cve_data[cve]['Exploit Status'].add(exploit_status)
                cve_data[cve]['Total Occurrences'] += 1
        
        if not cve_data:
            return
        
        # Convert to list format
        result = []
        for cve, data in sorted(cve_data.items()):
            affected_hosts_str = ', '.join(sorted(set(data['Affected Hosts'])))
            vulnerabilities_str = ', '.join(sorted(set(data['Vulnerabilities'])))
            severities_str = ', '.join(sorted(data['Severities']))
            exploit_status_str = ', '.join(sorted(data['Exploit Status'])) if data['Exploit Status'] else 'Unknown'
            
            avg_cvss = sum(data['CVSS Scores']) / len(data['CVSS Scores']) if data['CVSS Scores'] else 'N/A'
            max_cvss = max(data['CVSS Scores']) if data['CVSS Scores'] else 'N/A'
            
            result.append({
                'CVE': cve,
                'Total Occurrences': data['Total Occurrences'],
                'Affected Hosts Count': len(set(data['Affected Hosts'])),
                'Affected Hosts': affected_hosts_str,
                'Vulnerabilities': vulnerabilities_str,
                'Severities': severities_str,
                'Max CVSS Score': max_cvss,
                'Avg CVSS Score': f'{avg_cvss:.1f}' if isinstance(avg_cvss, float) else avg_cvss,
                'Exploit Status': exploit_status_str
            })
        
        # Sort by total occurrences (descending)
        cve_df = pd.DataFrame(result)
        cve_df = cve_df.sort_values('Total Occurrences', ascending=False)
        
        # Write to Excel
        cve_df.to_excel(writer, sheet_name='CVE Summary', index=False)
    
    def _create_pci_dss_sheet(self, writer: pd.ExcelWriter, df: pd.DataFrame) -> None:
        """Create PCI-DSS compliance sheet with pass/fail status."""
        # PCI-DSS requirements mapping
        pci_data = []
        
        for _, row in df.iterrows():
            vulnerability = row.get('Vulnerability', '')
            severity = row.get('Severity', '')
            cvss_score = row.get('CVSS Base Score', '')
            port = row.get('Port', '')
            protocol = row.get('Protocol', '')
            exposure = row.get('Exposure', '')
            exploit_status = row.get('Exploit Status', '')
            risk_score = row.get('Risk Score', 0)
            
            # Determine PCI-DSS compliance status
            pci_status = 'PASS'
            pci_reason = ''
            
            # Critical severity always fails
            if severity == 'Critical':
                pci_status = 'FAIL'
                pci_reason = 'Critical severity vulnerability'
            # High severity with external exposure fails
            elif severity == 'High' and exposure == 'External':
                pci_status = 'FAIL'
                pci_reason = 'High severity with external exposure'
            # High severity with exploit fails
            elif severity == 'High' and exploit_status and 'yes' in exploit_status.lower():
                pci_status = 'FAIL'
                pci_reason = 'High severity with available exploit'
            # CVSS >= 7.0 with external exposure fails
            elif cvss_score and cvss_score != 'N/A':
                try:
                    cvss_val = float(cvss_score)
                    if cvss_val >= 7.0 and exposure == 'External':
                        pci_status = 'FAIL'
                        pci_reason = f'CVSS {cvss_val:.1f} with external exposure'
                    elif cvss_val >= 9.0:
                        pci_status = 'FAIL'
                        pci_reason = f'Critical CVSS score ({cvss_val:.1f})'
                except (ValueError, TypeError):
                    pass
            # Risk score >= 80 fails
            elif risk_score >= 80:
                pci_status = 'FAIL'
                pci_reason = f'High risk score ({risk_score:.1f})'
            # High risk ports with external exposure
            elif exposure == 'External' and port and port != 'N/A':
                try:
                    port_num = int(port)
                    if port_num in [21, 22, 23, 25, 80, 143, 443, 445, 3389]:
                        pci_status = 'FAIL'
                        pci_reason = f'High-risk port {port_num} with external exposure'
                except (ValueError, TypeError):
                    pass
            
            if pci_status == 'PASS':
                pci_reason = 'Compliant'
            
            pci_data.append({
                'Vulnerability': vulnerability,
                'Host IP': row.get('Host IP', ''),
                'Exposure': exposure,
                'Severity': severity,
                'CVSS Score': cvss_score if cvss_score != 'N/A' else '',
                'Risk Score': risk_score,
                'Port': port if port != 'N/A' else '',
                'Protocol': protocol if protocol != 'N/A' else '',
                'Exploit Status': exploit_status,
                'PCI-DSS Status': pci_status,
                'Reason': pci_reason,
                'CVE': row.get('CVE', ''),
                'Description': row.get('Synopsis', '')[:100]  # Truncate for readability
            })
        
        if not pci_data:
            return
        
        # Create DataFrame
        pci_df = pd.DataFrame(pci_data)
        
        # Sort by PCI-DSS status (FAIL first) and then by risk score
        pci_df['Status Order'] = pci_df['PCI-DSS Status'].map({'FAIL': 0, 'PASS': 1})
        pci_df = pci_df.sort_values(['Status Order', 'Risk Score'], ascending=[True, False])
        pci_df = pci_df.drop('Status Order', axis=1)
        
        # Write to Excel
        pci_df.to_excel(writer, sheet_name='PCI-DSS Compliance', index=False)
    
    def _create_severity_sheets(self, writer: pd.ExcelWriter, df: pd.DataFrame) -> None:
        """Create separate sheets for each severity level with grouped vulnerabilities."""
        for severity in SEVERITY_LEVELS[:4]:  # Critical, High, Medium, Low
            severity_df = df[df['Severity'] == severity]
            
            if len(severity_df) == 0:
                if self.logger:
                    self.logger.debug(f"No {severity} vulnerabilities found, skipping sheet")
                continue
            
            # Group by plugin and aggregate affected hosts
            grouped_data = self._group_vulnerabilities_by_plugin(severity_df)
            
            # Create grouped DataFrame
            grouped_df = pd.DataFrame(grouped_data)
            
            # Sort by affected host count (descending) and then by vulnerability name
            grouped_df = grouped_df.sort_values(
                ['Affected Hosts Count', 'Vulnerability'],
                ascending=[False, True]
            )
            
            # Write to Excel
            grouped_df.to_excel(writer, sheet_name=severity, index=False)
            
            # Store plugin counts for this severity (for top plugins section)
            plugin_counts = Counter(severity_df['Vulnerability'])
            self._add_top_plugins_to_sheet(writer, severity, plugin_counts, len(severity_df))
    
    def _group_vulnerabilities_by_plugin(
        self, 
        severity_df: pd.DataFrame
    ) -> List[Dict[str, Any]]:
        """Group vulnerabilities by plugin and aggregate affected hosts."""
        grouped = defaultdict(lambda: {
            'Vulnerability': '',
            'Synopsis': '',
            'Description': '',
            'Solution': '',
            'CVSS Base Score': '',
            'CVSS Vector': '',
            'CVE': '',
            'MITRE ATT&CK': '',
            'Exploit Status': '',
            'Exploit Available': '',
            'Exploitability Ease': '',
            'Affected Hosts': [],
            'Affected Hosts Count': 0,
            'Ports': set(),
            'Services': set()
        })
        
        for _, row in severity_df.iterrows():
            vulnerability_name = row['Vulnerability']
            
            if grouped[vulnerability_name]['Vulnerability'] == '':
                grouped[vulnerability_name]['Vulnerability'] = vulnerability_name
                grouped[vulnerability_name]['Synopsis'] = row.get('Synopsis', '')
                grouped[vulnerability_name]['Description'] = row.get('Description', '')
                grouped[vulnerability_name]['Solution'] = row.get('Solution', '')
                grouped[vulnerability_name]['CVSS Base Score'] = row.get('CVSS Base Score', '')
                grouped[vulnerability_name]['CVSS Vector'] = row.get('CVSS Vector', '')
                grouped[vulnerability_name]['CVE'] = row.get('CVE', '')
                grouped[vulnerability_name]['MITRE ATT&CK'] = row.get('MITRE ATT&CK', '')
                grouped[vulnerability_name]['Exploit Status'] = row.get('Exploit Status', 'Unknown')
                grouped[vulnerability_name]['Exploit Available'] = row.get('Exploit Available', '')
                grouped[vulnerability_name]['Exploitability Ease'] = row.get('Exploitability Ease', '')
            
            # Add affected host
            host_info = f"{row['Host IP']}"
            if row.get('Host Name', '') and row['Host Name'] != row['Host IP']:
                host_info += f" ({row['Host Name']})"
            
            if row.get('Port', '') and row['Port'] != 'N/A':
                host_info += f":{row['Port']}"
                if row.get('Protocol', '') and row['Protocol'] != 'N/A':
                    host_info += f"/{row['Protocol']}"
            
            grouped[vulnerability_name]['Affected Hosts'].append(host_info)
            
            # Collect ports and services
            if row.get('Port', '') and row['Port'] != 'N/A':
                grouped[vulnerability_name]['Ports'].add(str(row['Port']))
            if row.get('Service', '') and row['Service'] != 'N/A':
                grouped[vulnerability_name]['Services'].add(str(row['Service']))
        
        # Convert to list format
        result = []
        for vulnerability_name, data in grouped.items():
            # Combine affected hosts
            affected_hosts_str = ', '.join(sorted(set(data['Affected Hosts'])))
            
            # Combine ports
            ports_str = ', '.join(sorted(data['Ports'])) if data['Ports'] else 'N/A'
            
            # Combine services
            services_str = ', '.join(sorted(data['Services'])) if data['Services'] else 'N/A'
            
            result.append({
                'Vulnerability': data['Vulnerability'],
                'Affected Hosts Count': len(set(data['Affected Hosts'])),
                'Affected Hosts': affected_hosts_str,
                'Ports': ports_str,
                'Services': services_str,
                'Exploit Status': data['Exploit Status'],
                'Synopsis': data['Synopsis'],
                'CVE': data['CVE'],
                'MITRE ATT&CK': data['MITRE ATT&CK'],
                'CVSS Base Score': data['CVSS Base Score'],
                'CVSS Vector': data['CVSS Vector'],
                'Description': data['Description'],
                'Solution': data['Solution']
            })
        
        return result
    
    def _add_top_plugins_to_sheet(
        self,
        writer: pd.ExcelWriter,
        severity: str,
        plugin_counts: Counter,
        total_count: int
    ) -> None:
        """Add top plugins section to severity sheet."""
        wb = writer.book
        ws = wb[severity]
        
        # Find the last row
        max_row = ws.max_row
        
        # Add spacing
        current_row = max_row + 3
        
        # Add title
        ws.cell(
            row=current_row,
            column=1,
            value=f'Top {TOP_PLUGINS_COUNT} Most Common {severity} Vulnerabilities'
        )
        ws.cell(row=current_row, column=1).font = Font(
            bold=True,
            size=SECTION_HEADER_FONT_SIZE,
            color=HEADER_COLOR
        )
        current_row += 1
        
        # Add headers
        ws.cell(row=current_row, column=1, value='Vulnerability')
        ws.cell(row=current_row, column=2, value='Occurrence Count')
        ws.cell(row=current_row, column=3, value='Percentage')
        
        header_fill, header_font = self.format_helper.get_header_style()
        
        for col in [1, 2, 3]:
            cell = ws.cell(row=current_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        current_row += 1
        plugins_start_row = current_row
        
        # Add top N plugins
        top_plugins = sorted(
            plugin_counts.items(),
            key=lambda x: x[1],
            reverse=True
        )[:TOP_PLUGINS_COUNT]
        
        for plugin_name, count in top_plugins:
            percentage = (count / total_count) * 100 if total_count > 0 else 0
            ws.cell(row=current_row, column=1, value=plugin_name)
            ws.cell(row=current_row, column=2, value=count)
            ws.cell(row=current_row, column=3, value=f'{percentage:.1f}%')
            current_row += 1
        
        plugins_end_row = current_row - 1
        
        # Store for chart creation
        self.severity_plugin_positions[severity] = {
            'start': plugins_start_row,
            'end': plugins_end_row
        }
    
    def _format_excel(self, file_path: Path) -> None:
        """Apply professional formatting to Excel file."""
        try:
            wb = load_workbook(file_path)
            
            # Format Summary sheet
            if 'Summary' in wb.sheetnames:
                self._format_summary_sheet(wb['Summary'])
            
            # Format CVE Summary sheet
            if 'CVE Summary' in wb.sheetnames:
                self._format_cve_sheet(wb['CVE Summary'])
            
            # Format PCI-DSS Compliance sheet
            if 'PCI-DSS Compliance' in wb.sheetnames:
                self._format_pci_dss_sheet(wb['PCI-DSS Compliance'])
            
            # Format severity-specific sheets
            for severity in SEVERITY_LEVELS[:4]:
                if severity in wb.sheetnames:
                    self._format_severity_sheet(wb[severity], severity)
            
            # Format All Vulnerabilities sheet
            if 'All Vulnerabilities' in wb.sheetnames:
                self._format_all_vulnerabilities_sheet(wb['All Vulnerabilities'])
            
            # Set Summary as first sheet
            if 'Summary' in wb.sheetnames:
                wb.move_sheet('Summary', offset=-len(wb.sheetnames))
            
            wb.save(file_path)
            if self.logger:
                self.logger.debug("Excel formatting completed")
            
        except Exception as e:
            if self.logger:
                self.logger.error(f"Error formatting Excel file: {e}", exc_info=True)
            raise
    
    def _format_summary_sheet(self, ws: Any) -> None:
        """Format Summary sheet."""
        title_fill, title_font = self.format_helper.get_title_style()
        header_fill, header_font = self.format_helper.get_header_style()
        thin_border = self.format_helper.get_border()
        
        # Format title row
        ws.cell(row=1, column=1).font = title_font
        ws.merge_cells('A1:B1')
        
        # Format all rows
        for row_idx, row in enumerate(ws.iter_rows(), 1):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Format section headers
                if cell.value and isinstance(cell.value, str):
                    if 'Severity Distribution' in cell.value or 'Top' in cell.value and 'Hosts' in cell.value:
                        cell.font = Font(
                            bold=True,
                            size=SECTION_HEADER_FONT_SIZE,
                            color=HEADER_COLOR
                        )
                        cell.fill = title_fill
                
                # Format table headers
                if row_idx in [self.summary_severity_start - 1, self.summary_hosts_start - 1] and cell.value:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        self._auto_adjust_column_widths(ws, max_width=50)
        
        # Add charts to Summary sheet
        self._add_summary_charts(ws)
    
    def _format_cve_sheet(self, ws: Any) -> None:
        """Format CVE Summary sheet."""
        header_fill, header_font = self.format_helper.get_header_style()
        fill_even, fill_odd = self.format_helper.get_row_fills()
        thin_border = self.format_helper.get_border()
        
        # Format header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )
            cell.border = thin_border
        
        # Find column indices
        header_row = ws[1]
        cve_col = None
        affected_hosts_col = None
        affected_hosts_count_col = None
        max_cvss_col = None
        
        for idx, cell in enumerate(header_row, 1):
            if cell.value == 'CVE':
                cve_col = idx
            elif cell.value == 'Affected Hosts':
                affected_hosts_col = idx
            elif cell.value == 'Affected Hosts Count':
                affected_hosts_count_col = idx
            elif cell.value == 'Max CVSS Score':
                max_cvss_col = idx
        
        # Format data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            fill_color = fill_even if row_idx % 2 == 0 else fill_odd
            for col_idx, cell in enumerate(row, 1):
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal='left',
                    vertical='top',
                    wrap_text=True
                )
                cell.fill = fill_color
                
                # Format CVE column
                if col_idx == cve_col:
                    cell.font = Font(bold=True, size=11)
                    cell.alignment = Alignment(
                        horizontal='left',
                        vertical='center'
                    )
                
                # Format Affected Hosts column
                elif col_idx == affected_hosts_col:
                    cell.font = Font(size=SMALL_FONT_SIZE)
                
                # Format Affected Hosts Count column
                elif col_idx == affected_hosts_count_col:
                    cell.alignment = Alignment(
                        horizontal='center',
                        vertical='center'
                    )
                    cell.font = Font(bold=True, size=DATA_FONT_SIZE)
                
                # Format Max CVSS Score column
                elif col_idx == max_cvss_col:
                    cell.alignment = Alignment(
                        horizontal='center',
                        vertical='center'
                    )
                    if cell.value and isinstance(cell.value, (int, float)):
                        score = float(cell.value)
                        if score >= 9.0:
                            cell.fill = PatternFill(
                                start_color='FF0000',
                                end_color='FF0000',
                                fill_type="solid"
                            )
                            cell.font = Font(bold=True, color='FFFFFF', size=11)
                        elif score >= 7.0:
                            cell.fill = PatternFill(
                                start_color='FF6600',
                                end_color='FF6600',
                                fill_type="solid"
                            )
                            cell.font = Font(bold=True, color='FFFFFF', size=11)
                        elif score >= 4.0:
                            cell.fill = PatternFill(
                                start_color='FFCC00',
                                end_color='FFCC00',
                                fill_type="solid"
                            )
                            cell.font = Font(bold=True, size=11)
        
        # Auto-adjust column widths
        self._auto_adjust_column_widths(ws, max_width=MAX_COLUMN_WIDTH)
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _format_pci_dss_sheet(self, ws: Any) -> None:
        """Format PCI-DSS Compliance sheet."""
        header_fill, header_font = self.format_helper.get_header_style()
        fill_even, fill_odd = self.format_helper.get_row_fills()
        thin_border = self.format_helper.get_border()
        
        # Format header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )
            cell.border = thin_border
        
        # Find column indices
        header_row = ws[1]
        pci_status_col = None
        severity_col = None
        risk_score_col = None
        
        for idx, cell in enumerate(header_row, 1):
            if cell.value == 'PCI-DSS Status':
                pci_status_col = idx
            elif cell.value == 'Severity':
                severity_col = idx
            elif cell.value == 'Risk Score':
                risk_score_col = idx
        
        # Format data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            fill_color = fill_even if row_idx % 2 == 0 else fill_odd
            for col_idx, cell in enumerate(row, 1):
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal='left',
                    vertical='top',
                    wrap_text=True
                )
                cell.fill = fill_color
                
                # Format PCI-DSS Status column with color coding
                if col_idx == pci_status_col:
                    cell.alignment = Alignment(
                        horizontal='center',
                        vertical='center'
                    )
                    cell.font = Font(bold=True, size=11)
                    if cell.value == 'FAIL':
                        cell.fill = PatternFill(
                            start_color='FF0000',
                            end_color='FF0000',
                            fill_type="solid"
                        )
                        cell.font = Font(bold=True, color='FFFFFF', size=11)
                    elif cell.value == 'PASS':
                        cell.fill = PatternFill(
                            start_color='00CC00',
                            end_color='00CC00',
                            fill_type="solid"
                        )
                        cell.font = Font(bold=True, color='FFFFFF', size=11)
                
                # Format Severity column
                elif col_idx == severity_col:
                    severity = cell.value
                    severity_fill = self.format_helper.get_severity_fill(severity)
                    if severity_fill:
                        cell.fill = severity_fill
                        if severity in ['Critical', 'High']:
                            cell.font = Font(bold=True, color='FFFFFF')
                
                # Format Risk Score column
                elif col_idx == risk_score_col:
                    cell.alignment = Alignment(
                        horizontal='center',
                        vertical='center'
                    )
                    if cell.value:
                        try:
                            score = float(cell.value) if isinstance(cell.value, (int, float)) else 0
                            if score >= 90:
                                cell.fill = PatternFill(
                                    start_color='FF0000',
                                    end_color='FF0000',
                                    fill_type="solid"
                                )
                                cell.font = Font(bold=True, color='FFFFFF', size=11)
                            elif score >= 70:
                                cell.fill = PatternFill(
                                    start_color='FF6600',
                                    end_color='FF6600',
                                    fill_type="solid"
                                )
                                cell.font = Font(bold=True, color='FFFFFF', size=11)
                            elif score >= 50:
                                cell.fill = PatternFill(
                                    start_color='FFCC00',
                                    end_color='FFCC00',
                                    fill_type="solid"
                                )
                                cell.font = Font(bold=True, size=11)
                        except (ValueError, TypeError):
                            pass
        
        # Auto-adjust column widths
        self._auto_adjust_column_widths(ws, max_width=MAX_COLUMN_WIDTH)
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _format_all_vulnerabilities_sheet(self, ws: Any) -> None:
        """Format All Vulnerabilities sheet."""
        header_fill, header_font = self.format_helper.get_header_style()
        fill_even, fill_odd = self.format_helper.get_row_fills()
        thin_border = self.format_helper.get_border()
        
        # Format header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )
            cell.border = thin_border
        
        # Find Exploit Status column index
        exploit_status_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'Exploit Status':
                exploit_status_col = idx
                break
        
        # Format data rows with alternating colors
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            fill_color = fill_even if row_idx % 2 == 0 else fill_odd
            for col_idx, cell in enumerate(row, 1):
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal='left',
                    vertical='top',
                    wrap_text=True
                )
                cell.fill = fill_color
                
                # Color code severity column
                if cell.column == 5:  # Severity column
                    severity = cell.value
                    severity_fill = self.format_helper.get_severity_fill(severity)
                    if severity_fill:
                        cell.fill = severity_fill
                        if severity in ['Critical', 'High']:
                            cell.font = Font(bold=True, color='FFFFFF')
                
                # Format Exploit Status column
                elif exploit_status_col and col_idx == exploit_status_col:
                    cell.alignment = Alignment(
                        horizontal='center',
                        vertical='center'
                    )
                    # Color code based on exploit status
                    if cell.value:
                        exploit_status = str(cell.value).strip().lower()
                        if exploit_status.startswith('yes'):
                            cell.fill = PatternFill(
                                start_color='FF0000',
                                end_color='FF0000',
                                fill_type="solid"
                            )
                            cell.font = Font(bold=True, color='FFFFFF', size=10)
                        elif 'possibly' in exploit_status:
                            cell.fill = PatternFill(
                                start_color='FF9900',
                                end_color='FF9900',
                                fill_type="solid"
                            )
                            cell.font = Font(bold=True, color='FFFFFF', size=10)
                        elif exploit_status == 'no':
                            cell.fill = PatternFill(
                                start_color='00CC00',
                                end_color='00CC00',
                                fill_type="solid"
                            )
                            cell.font = Font(bold=True, color='FFFFFF', size=10)
                        else:
                            # Unknown - keep default fill but bold font
                            cell.font = Font(bold=True, size=10)
                    else:
                        cell.font = Font(bold=True, size=10)
        
        # Auto-adjust column widths
        self._auto_adjust_column_widths(ws, max_width=MAX_COLUMN_WIDTH)
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _format_cve_sheet(self, ws: Any) -> None:
        """Format CVE Summary sheet."""
        header_fill, header_font = self.format_helper.get_header_style()
        fill_even, fill_odd = self.format_helper.get_row_fills()
        thin_border = self.format_helper.get_border()
        
        # Format header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )
            cell.border = thin_border
        
        # Find column indices
        header_row = ws[1]
        cve_col = None
        affected_hosts_col = None
        affected_hosts_count_col = None
        max_cvss_col = None
        
        for idx, cell in enumerate(header_row, 1):
            if cell.value == 'CVE':
                cve_col = idx
            elif cell.value == 'Affected Hosts':
                affected_hosts_col = idx
            elif cell.value == 'Affected Hosts Count':
                affected_hosts_count_col = idx
            elif cell.value == 'Max CVSS Score':
                max_cvss_col = idx
        
        # Format data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            fill_color = fill_even if row_idx % 2 == 0 else fill_odd
            for col_idx, cell in enumerate(row, 1):
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal='left',
                    vertical='top',
                    wrap_text=True
                )
                cell.fill = fill_color
                
                # Format CVE column
                if col_idx == cve_col:
                    cell.font = Font(bold=True, size=11)
                    cell.alignment = Alignment(
                        horizontal='left',
                        vertical='center'
                    )
                
                # Format Affected Hosts column
                elif col_idx == affected_hosts_col:
                    cell.font = Font(size=SMALL_FONT_SIZE)
                
                # Format Affected Hosts Count column
                elif col_idx == affected_hosts_count_col:
                    cell.alignment = Alignment(
                        horizontal='center',
                        vertical='center'
                    )
                    cell.font = Font(bold=True, size=DATA_FONT_SIZE)
                
                # Format Max CVSS Score column
                elif col_idx == max_cvss_col:
                    cell.alignment = Alignment(
                        horizontal='center',
                        vertical='center'
                    )
                    if cell.value and isinstance(cell.value, (int, float)):
                        score = float(cell.value)
                        if score >= 9.0:
                            cell.fill = PatternFill(
                                start_color='FF0000',
                                end_color='FF0000',
                                fill_type="solid"
                            )
                            cell.font = Font(bold=True, color='FFFFFF', size=11)
                        elif score >= 7.0:
                            cell.fill = PatternFill(
                                start_color='FF6600',
                                end_color='FF6600',
                                fill_type="solid"
                            )
                            cell.font = Font(bold=True, color='FFFFFF', size=11)
                        elif score >= 4.0:
                            cell.fill = PatternFill(
                                start_color='FFCC00',
                                end_color='FFCC00',
                                fill_type="solid"
                            )
                            cell.font = Font(bold=True, size=11)
        
        # Auto-adjust column widths
        self._auto_adjust_column_widths(ws, max_width=MAX_COLUMN_WIDTH)
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _format_severity_sheet(self, ws: Any, severity: str) -> None:
        """Format a severity-specific sheet."""
        header_fill, header_font = self.format_helper.get_header_style()
        severity_color = SEVERITY_COLORS.get(severity, HEADER_COLOR)
        severity_fill = PatternFill(
            start_color=severity_color,
            end_color=severity_color,
            fill_type="solid"
        )
        fill_even, fill_odd = self.format_helper.get_row_fills()
        thin_border = self.format_helper.get_border()
        
        # Format header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = severity_fill
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )
            cell.border = thin_border
        
        # Find the row where top plugins section starts
        top_plugins_start = None
        for row_idx, row in enumerate(ws.iter_rows(), 1):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and 'Top' in cell.value and 'Common' in cell.value:
                    top_plugins_start = row_idx
                    break
            if top_plugins_start:
                break
        
        # Format vulnerability data rows
        max_data_row = top_plugins_start - 3 if top_plugins_start else ws.max_row
        
        # Find column indices for special formatting
        header_row = ws[1]
        affected_hosts_col = None
        affected_hosts_count_col = None
        vulnerability_col = None
        exploit_status_col = None
        
        for idx, cell in enumerate(header_row, 1):
            if cell.value == 'Affected Hosts':
                affected_hosts_col = idx
            elif cell.value == 'Affected Hosts Count':
                affected_hosts_count_col = idx
            elif cell.value == 'Vulnerability':
                vulnerability_col = idx
            elif cell.value == 'Exploit Status':
                exploit_status_col = idx
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=max_data_row), 2):
            fill_color = fill_even if row_idx % 2 == 0 else fill_odd
            for col_idx, cell in enumerate(row, 1):
                cell.border = thin_border
                cell.fill = fill_color
                
                # Special formatting for Affected Hosts column
                if col_idx == affected_hosts_col:
                    cell.alignment = Alignment(
                        horizontal='left',
                        vertical='top',
                        wrap_text=True
                    )
                    cell.font = Font(size=SMALL_FONT_SIZE)
                
                # Special formatting for Affected Hosts Count column
                elif col_idx == affected_hosts_count_col:
                    cell.alignment = Alignment(
                        horizontal='center',
                        vertical='center'
                    )
                    cell.font = Font(bold=True, size=DATA_FONT_SIZE)
                    # Color code based on count
                    if cell.value:
                        try:
                            count = int(cell.value) if isinstance(cell.value, (int, float)) else 0
                            if count >= 10:
                                cell.fill = PatternFill(
                                    start_color=HIGH_COUNT_COLOR,
                                    end_color=HIGH_COUNT_COLOR,
                                    fill_type="solid"
                                )
                            elif count >= 5:
                                cell.fill = PatternFill(
                                    start_color=MEDIUM_COUNT_COLOR,
                                    end_color=MEDIUM_COUNT_COLOR,
                                    fill_type="solid"
                                )
                        except (ValueError, TypeError):
                            pass
                
                # Format Vulnerability column
                elif col_idx == vulnerability_col:
                    cell.alignment = Alignment(
                        horizontal='left',
                        vertical='top',
                        wrap_text=True
                    )
                    cell.font = Font(bold=True, size=10)
                
                # Format Exploit Status column
                elif exploit_status_col and col_idx == exploit_status_col:
                    cell.alignment = Alignment(
                        horizontal='center',
                        vertical='center'
                    )
                    # Color code based on exploit status
                    if cell.value:
                        exploit_status = str(cell.value).strip().lower()
                        if exploit_status.startswith('yes'):
                            cell.fill = PatternFill(
                                start_color='FF0000',
                                end_color='FF0000',
                                fill_type="solid"
                            )
                            cell.font = Font(bold=True, color='FFFFFF', size=10)
                        elif 'possibly' in exploit_status:
                            cell.fill = PatternFill(
                                start_color='FF9900',
                                end_color='FF9900',
                                fill_type="solid"
                            )
                            cell.font = Font(bold=True, color='FFFFFF', size=10)
                        elif exploit_status == 'no':
                            cell.fill = PatternFill(
                                start_color='00CC00',
                                end_color='00CC00',
                                fill_type="solid"
                            )
                            cell.font = Font(bold=True, color='FFFFFF', size=10)
                        else:
                            # Unknown - keep default fill but bold font
                            cell.font = Font(bold=True, size=10)
                    else:
                        cell.font = Font(bold=True, size=10)
                else:
                    cell.alignment = Alignment(
                        horizontal='left',
                        vertical='top',
                        wrap_text=True
                    )
        
        # Format top plugins section
        if top_plugins_start and severity in self.severity_plugin_positions:
            pos = self.severity_plugin_positions[severity]
            
            # Format title
            title_cell = ws.cell(row=top_plugins_start, column=1)
            title_cell.font = Font(
                bold=True,
                size=SECTION_HEADER_FONT_SIZE,
                color=severity_color
            )
            ws.merge_cells(f'A{top_plugins_start}:C{top_plugins_start}')
            
            # Format header row
            header_row = top_plugins_start + 1
            for col in [1, 2, 3]:
                cell = ws.cell(row=header_row, column=col)
                cell.font = header_font
                cell.fill = severity_fill
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center'
                )
                cell.border = thin_border
            
            # Format data rows
            for row_idx in range(pos['start'], pos['end'] + 1):
                fill_color = fill_even if row_idx % 2 == 0 else fill_odd
                for col in [1, 2, 3]:
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = thin_border
                    cell.alignment = Alignment(
                        horizontal='left',
                        vertical='center'
                    )
                    cell.fill = fill_color
            
            # Add bar chart for top plugins
            self._add_severity_chart(ws, severity, pos, top_plugins_start)
        
        # Auto-adjust column widths with special handling
        self._auto_adjust_severity_column_widths(ws)
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _add_severity_chart(
        self,
        ws: Any,
        severity: str,
        pos: Dict[str, int],
        chart_start_row: int
    ) -> None:
        """Add bar chart to severity sheet."""
        bar = BarChart()
        bar.type = "col"
        bar.style = 10
        bar.title = f"Top {TOP_PLUGINS_COUNT} Most Common {severity} Vulnerabilities"
        bar.y_axis.title = "Occurrence Count"
        bar.x_axis.title = "Vulnerability"
        bar.height = CHART_HEIGHT
        bar.width = CHART_WIDTH
        
        labels = Reference(
            ws,
            min_col=1,
            min_row=pos['start'],
            max_row=pos['end']
        )
        data = Reference(
            ws,
            min_col=2,
            min_row=pos['start'] - 1,
            max_row=pos['end']
        )
        
        bar.add_data(data, titles_from_data=False)
        bar.set_categories(labels)
        bar.dataLabels = DataLabelList()
        bar.dataLabels.showVal = True
        
        # Format axis labels
        bar.x_axis.tickLblPos = 'low'
        bar.x_axis.majorTickMark = 'out'
        bar.y_axis.majorTickMark = 'out'
        bar.x_axis.minorTickMark = 'none'
        bar.y_axis.minorTickMark = 'none'
        
        # Format data labels
        if hasattr(bar.dataLabels, 'font'):
            bar.dataLabels.font.size = CHART_LABEL_FONT_SIZE
            bar.dataLabels.font.bold = True
        
        # Format legend
        bar.legend.position = 'r'
        if hasattr(bar.legend, 'font'):
            bar.legend.font.size = CHART_LABEL_FONT_SIZE
        
        # Place chart to the right of the table
        ws.add_chart(bar, f"E{chart_start_row}")
    
    def _add_summary_charts(self, ws: Any) -> None:
        """Add charts to Summary sheet with improved formatting."""
        # Pie chart for severity distribution
        pie = PieChart()
        pie.title = "Vulnerability Distribution by Severity"
        pie.height = CHART_HEIGHT
        pie.width = CHART_WIDTH
        
        labels = Reference(
            ws,
            min_col=1,
            min_row=self.summary_severity_start,
            max_row=self.summary_severity_end
        )
        data = Reference(
            ws,
            min_col=2,
            min_row=self.summary_severity_start - 1,
            max_row=self.summary_severity_end
        )
        
        pie.add_data(data, titles_from_data=False)
        pie.set_categories(labels)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showLeaderLines = True
        
        # Format legend - position at bottom with better spacing
        pie.legend.position = 'b'
        if hasattr(pie.legend, 'font'):
            pie.legend.font.size = 11
        
        # Format data labels
        if hasattr(pie.dataLabels, 'font'):
            pie.dataLabels.font.size = 10
            pie.dataLabels.font.bold = True
        # Set label position to bestFit for automatic positioning
        if hasattr(pie.dataLabels, 'dLblPos'):
            pie.dataLabels.dLblPos = 'bestFit'
        
        ws.add_chart(pie, "D2")
        
        # Bar chart for top hosts with improved formatting
        bar = BarChart()
        bar.type = "col"
        bar.style = 10
        bar.title = f"Top {TOP_HOSTS_COUNT} Most Affected Hosts"
        bar.y_axis.title = "Vulnerability Count"
        bar.x_axis.title = "Host IP"
        bar.height = CHART_HEIGHT
        bar.width = CHART_WIDTH
        
        labels_hosts = Reference(
            ws,
            min_col=1,
            min_row=self.summary_hosts_start,
            max_row=self.summary_hosts_end
        )
        data_hosts = Reference(
            ws,
            min_col=2,
            min_row=self.summary_hosts_start - 1,
            max_row=self.summary_hosts_end
        )
        
        bar.add_data(data_hosts, titles_from_data=False)
        bar.set_categories(labels_hosts)
        bar.dataLabels = DataLabelList()
        bar.dataLabels.showVal = True
        
        # Format axis labels - prevent overlap
        bar.x_axis.majorTickMark = 'out'
        bar.y_axis.majorTickMark = 'out'
        bar.x_axis.minorTickMark = 'none'
        bar.y_axis.minorTickMark = 'none'
        bar.x_axis.tickLblPos = 'low'  # Position labels below
        
        # Format data labels
        if hasattr(bar.dataLabels, 'font'):
            bar.dataLabels.font.size = 9
            bar.dataLabels.font.bold = True
        # Set label position to outEnd for labels outside bars
        if hasattr(bar.dataLabels, 'dLblPos'):
            bar.dataLabels.dLblPos = 'outEnd'
        
        # Format legend
        bar.legend.position = 'r'
        if hasattr(bar.legend, 'font'):
            bar.legend.font.size = 11
        
        ws.add_chart(bar, "D25")
        
        # Pie chart for exposure distribution
        if hasattr(self, 'summary_exposure_start') and hasattr(self, 'summary_exposure_end'):
            exposure_pie = PieChart()
            exposure_pie.title = "Exposure Distribution"
            exposure_pie.height = CHART_HEIGHT
            exposure_pie.width = CHART_WIDTH
            
            labels_exposure = Reference(
                ws,
                min_col=1,
                min_row=self.summary_exposure_start,
                max_row=self.summary_exposure_end
            )
            data_exposure = Reference(
                ws,
                min_col=2,
                min_row=self.summary_exposure_start - 1,
                max_row=self.summary_exposure_end
            )
            
            exposure_pie.add_data(data_exposure, titles_from_data=False)
            exposure_pie.set_categories(labels_exposure)
            exposure_pie.dataLabels = DataLabelList()
            exposure_pie.dataLabels.showPercent = True
            exposure_pie.dataLabels.showLeaderLines = True
            
            exposure_pie.legend.position = 'b'
            if hasattr(exposure_pie.legend, 'font'):
                exposure_pie.legend.font.size = 11
            
            if hasattr(exposure_pie.dataLabels, 'font'):
                exposure_pie.dataLabels.font.size = 10
                exposure_pie.dataLabels.font.bold = True
            if hasattr(exposure_pie.dataLabels, 'dLblPos'):
                exposure_pie.dataLabels.dLblPos = 'bestFit'
            
            ws.add_chart(exposure_pie, "D42")
    
    def _auto_adjust_column_widths(self, ws: Any, max_width: int = MAX_COLUMN_WIDTH) -> None:
        """Auto-adjust column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, max_width)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _auto_adjust_severity_column_widths(self, ws: Any) -> None:
        """Auto-adjust column widths for severity sheets with special handling."""
        for col_idx, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            # Check header for column name
            header_cell = ws.cell(row=1, column=col_idx)
            column_name = header_cell.value if header_cell.value else ''
            
            for cell in column:
                try:
                    if cell.value:
                        if column_name == 'Affected Hosts':
                            max_length = max(max_length, AFFECTED_HOSTS_MIN_WIDTH)
                        else:
                            max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            
            # Set column widths
            if column_name == 'Affected Hosts':
                adjusted_width = max(max_length + 2, AFFECTED_HOSTS_MIN_WIDTH)
            elif column_name == 'Affected Hosts Count':
                adjusted_width = AFFECTED_HOSTS_COUNT_WIDTH
            elif column_name == 'Exploit Status':
                adjusted_width = 18
            elif column_name == 'Vulnerability':
                adjusted_width = min(max_length + 2, VULNERABILITY_MAX_WIDTH)
            elif column_name in ['Description', 'Solution']:
                adjusted_width = min(max_length + 2, DESCRIPTION_MAX_WIDTH)
            else:
                adjusted_width = min(max_length + 2, MAX_COLUMN_WIDTH)
            
            ws.column_dimensions[column_letter].width = adjusted_width


def main() -> None:
    """Main entry point for the CLI application."""
    parser = argparse.ArgumentParser(
        description='Convert Nessus XML scan results to professional Excel reports',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  %(prog)s scan1.nessus scan2.nessus -o report.xlsx
  %(prog)s *.nessus -o output.xlsx
  %(prog)s scan.nessus -o report.xlsx --verbose

Developed by: Cuma KURT
Email: cumakurt@gmail.com
LinkedIn: https://www.linkedin.com/in/cuma-kurt-34414917/

License: GNU General Public License v3.0 (GPL-3.0)
        '''
    )
    
    parser.add_argument(
        'input_files',
        nargs='+',
        type=str,
        help='One or more .nessus XML files to process'
    )
    
    parser.add_argument(
        '-o', '--output',
        type=str,
        default='nessus_report.xlsx',
        help='Output Excel file path (default: nessus_report.xlsx)'
    )
    
    parser.add_argument(
        '-v', '--verbose',
        action='store_true',
        help='Enable verbose logging'
    )
    
    parser.add_argument(
        '--log-file',
        type=str,
        default=None,
        help='Path to log file (optional)'
    )
    
    parser.add_argument(
        '--no-progress',
        action='store_true',
        help='Disable progress bar'
    )
    
    parser.add_argument(
        '--include-info',
        action='store_true',
        help='Include Info (severity 0) vulnerabilities in the report'
    )
    
    args = parser.parse_args()
    
    # Setup logger
    log_file = Path(args.log_file) if args.log_file else None
    logger = setup_logger(verbose=args.verbose, log_file=log_file)
    
    # Validate input files
    input_paths = []
    for file_path in args.input_files:
        path = Path(file_path)
        if not path.exists():
            print(f"Error: File not found: {file_path}")
            sys.exit(1)
        if path.suffix.lower() != '.nessus' and not args.verbose:
            pass  # Silent warning unless verbose
        input_paths.append(path)
    
    # Parse Nessus files
    print(f"Parsing {len(input_paths)} Nessus file(s)...")
    nessus_parser = NessusParser(
        logger=logger if args.verbose else None,
        include_info=args.include_info
    )
    
    # Process files with progress bar
    file_iterator = tqdm(
        input_paths,
        desc="Processing",
        disable=args.no_progress,
        leave=False
    )
    
    for file_path in file_iterator:
        nessus_parser.parse_file(file_path)
    
    # Report results
    if nessus_parser.failed_files:
        print(f"\nWarning: Failed to process {len(nessus_parser.failed_files)} file(s)")
        if args.verbose:
            for failed_file, error in nessus_parser.failed_files:
                logger.warning(f"  - {failed_file}: {error}")
    
    if not nessus_parser.vulnerabilities:
        print("Error: No vulnerabilities found in the input files.")
        sys.exit(1)
    
    total_vulns = len(nessus_parser.vulnerabilities)
    total_hosts = len(nessus_parser.host_vuln_count)
    print(f"Found {total_vulns:,} vulnerabilities across {total_hosts:,} hosts")
    
    # Generate Excel report
    output_path = Path(args.output)
    print(f"Generating Excel report: {output_path.name}")
    
    try:
        generator = ExcelReportGenerator(
            nessus_parser.vulnerabilities,
            nessus_parser.host_vuln_count,
            logger=logger if args.verbose else None
        )
        generator.generate(output_path)
        print(f"✓ Success! Report generated: {output_path}")
        
    except ExcelGenerationError as e:
        print(f"Error: Failed to generate Excel report: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"Error: Unexpected error occurred: {e}")
        if args.verbose:
            logger.error("Unexpected error", exc_info=True)
        sys.exit(1)


if __name__ == '__main__':
    main()
